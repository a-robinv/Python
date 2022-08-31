import openpyxl as xl
#shortens openpyxl ~ xl
from openpyxl.chart import BarChart, Reference


#loads xl and return workbook object
wb = xl.load_workbook('transactions.xlsx')
#access sheet
sheet = wb['Sheet1']
#get shell
cell = sheet['a1']
#or
cell = sheet.cell(1,1)
# print(cell.value)

for row in range (2, sheet.max_row + 1):
    cell = sheet.cell(row,3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value = corrected_price
#getting values for chartt
values = Reference(sheet,
          min_row=2 ,
          max_row = sheet.max_row,
          min_col=4,
          max_col=4)

#creating chart
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')

wb.save('transactions2.xlsx')